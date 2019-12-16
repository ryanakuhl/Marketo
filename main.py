from openpyxl import load_workbook
import datetime, xlsxwriter, xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook

a = []
def xls_to_xlsx(file):
    xlsBook = xlrd.open_workbook(file)
    workbook = openpyxlWorkbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)
    workbook.save(file.split('.')[0] + '.xlsx')
    workbook.close()


def write_to_excel(d, workbook):
    cell_format = workbook.add_format({'bold': True})
    worksheet = workbook.add_worksheet(str(d))
    worksheet.autofilter('A4:K4')
    worksheet.write(3, 0, 'Person Name', cell_format)
    worksheet.write(3, 1, 'Email', cell_format)
    worksheet.write(3, 2, 'Company', cell_format)
    worksheet.write(3, 3, 'Send Email', cell_format)
    worksheet.write(3, 4, 'Email Delivered', cell_format)
    worksheet.write(3, 5, 'Email Bounced Soft', cell_format)
    worksheet.write(3, 6, 'Email Bounced', cell_format)
    worksheet.write(3, 7, 'Open Email', cell_format)
    worksheet.write(3, 8, 'Visit Web Page', cell_format)
    worksheet.write(3, 9, 'Click Email', cell_format)
    worksheet.write(3, 10, 'Unsubscribe Email', cell_format)
    row = 4
    total_data = {'Send Email': 0,
                  'Email Delivered': 0,
                  'Email Bounced Soft': 0,
                  'Email Bounced': 0,
                  'Open Email': 0,
                  'Visit Web Page': 0,
                  'Click Email': 0,
                  'Unsubscribe Email': 0,
                  }
    for b in a:
        if b.Date == d:
            worksheet.write(row, 0, b.Person_Name)
            worksheet.write(row, 1, b.Email)
            worksheet.write(row, 2, b.Company)
            worksheet.write(row, 3, b.Activity.get('Send Email'))
            worksheet.write(row, 4, b.Activity.get('Email Delivered'))
            worksheet.write(row, 5, b.Activity.get('Email Bounced Soft'))
            worksheet.write(row, 6, b.Activity.get('Email Bounced'))
            worksheet.write(row, 7, b.Activity.get('Open Email'))
            worksheet.write(row, 8, b.Activity.get('Visit Web Page'))
            worksheet.write(row, 9, b.Activity.get('Click Email'))
            worksheet.write(row, 10, b.Activity.get('Unsubscribe Email'))
            for k in b.Activity:
                if bool(b.Activity.get(k)):
                    total_data[k] += 1
            row += 1

    worksheet.write(1, 3, 'Send Email', cell_format)
    worksheet.write(1, 4, 'Email Delivered', cell_format)
    worksheet.write(1, 5, 'Email Bounced Soft', cell_format)
    worksheet.write(1, 6, 'Email Bounced', cell_format)
    worksheet.write(1, 7, 'Open Email', cell_format)
    worksheet.write(1, 8, 'Visit Web Page', cell_format)
    worksheet.write(1, 9, 'Click Email', cell_format)
    worksheet.write(1, 10, 'Unsubscribe Email', cell_format)
    worksheet.write(2, 3, total_data.get('Send Email'))
    worksheet.write(2, 4, total_data.get('Email Delivered'))
    worksheet.write(2, 5, total_data.get('Email Bounced Soft'))
    worksheet.write(2, 6, total_data.get('Email Bounced'))
    worksheet.write(2, 7, total_data.get('Open Email'))
    worksheet.write(2, 8, total_data.get('Visit Web Page'))
    worksheet.write(2, 9, total_data.get('Click Email'))
    worksheet.write(2, 10, total_data.get('Unsubscribe Email'))


class FirstTab:
    def __init__(self, x):
        self.Event_ID = x[0].value
        self.Date = datetime.datetime.strptime(x[1].value, '%Y-%m-%d %H:%M:%S').date()
        self.Detail = x[3].value
        self.Person_ID = x[4].value
        self.Person_Name = x[5].value
        self.Email = x[6].value
        self.Company = x[7].value
        self.Activity = {'Send Email': 0,
                         'Email Delivered': 0,
                         'Email Bounced Soft': 0,
                         'Email Bounced': 0,
                         'Open Email': 0,
                         'Visit Web Page': 0,
                         'Click Email': 0,
                         'Unsubscribe Email': 0,
                         }


def get_overview(dates, workbook):
    cell_format = workbook.add_format({'bold': True})
    overview_stats = []
    for d in dates:
        total_data = {'Send Email': 0,
                      'Email Delivered': 0,
                      'Email Bounced Soft': 0,
                      'Email Bounced': 0,
                      'Open Email': 0,
                      'Visit Web Page': 0,
                      'Click Email': 0,
                      'Unsubscribe Email': 0,
                      }
        for b in a:
            if b.Date == d:
                for k in b.Activity:
                    if bool(b.Activity.get(k)):
                        total_data[k] += 1
        overview_stats.append([d, total_data])
    worksheet = workbook.add_worksheet('Overview')
    worksheet.autofilter('B2:I2')
    worksheet.write(1, 0, 'Send Date', cell_format)
    worksheet.write(1, 1, 'Send Email', cell_format)
    worksheet.write(1, 2, 'Email Delivered', cell_format)
    worksheet.write(1, 3, 'Email Bounced Soft', cell_format)
    worksheet.write(1, 4, 'Email Bounced', cell_format)
    worksheet.write(1, 5, 'Open Email', cell_format)
    worksheet.write(1, 6, 'Visit Web Page', cell_format)
    worksheet.write(1, 7, 'Click Email', cell_format)
    worksheet.write(1, 8, 'Unsubscribe Email', cell_format)
    row = 2
    for d in overview_stats:
        worksheet.write(row, 0, str(d[0]))
        worksheet.write(row, 1, d[1].get('Send Email'))
        worksheet.write(row, 2, d[1].get('Email Delivered'))
        worksheet.write(row, 3, d[1].get('Email Bounced Soft'))
        worksheet.write(row, 4, d[1].get('Email Bounced'))
        worksheet.write(row, 5, d[1].get('Open Email'))
        worksheet.write(row, 6, d[1].get('Visit Web Page'))
        worksheet.write(row, 7, d[1].get('Click Email'))
        worksheet.write(row, 8, d[1].get('Unsubscribe Email'))
        row += 1


def run(filename):
    if filename.split('.')[1] == 'xls':
        xls_to_xlsx(filename)
        filename = filename.split('.')[0] + '.xlsx'
    wb = load_workbook(filename=filename)
    s = wb.sheetnames[0]
    sheet_ranges = wb[s]
    email_addresses = []
    for row in sheet_ranges.rows:
        if row[2].value == 'Send Email':
            if row[5].value not in email_addresses:
                try:
                    x = FirstTab(row)
                    a.append(x)
                except Exception as e:
                    print(e)
                email_addresses.append(row[5].value)
    for row in sheet_ranges.rows:
        for b in a:
            if b.Email == row[6].value:
                b.Activity[row[2].value] += 1
    dates = []
    [dates.append(b.Date) for b in a if b.Date not in dates]
    workbook = xlsxwriter.Workbook(filename)
    get_overview(dates, workbook)
    for d in dates:
        write_to_excel(d, workbook)
    workbook.close()
